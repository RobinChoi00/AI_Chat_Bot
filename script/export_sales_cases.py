#!/usr/bin/env python3
"""Export Sales practical-case workbooks to gzipped CSVs under data/sales/.

Usage:
  python script/export_sales_cases.py \\
    --osaki raw_data/sales/osaki_practical_cases.xlsx \\
    --titan raw_data/sales/titan_practical_cases.xlsx
"""

from __future__ import annotations

import argparse
import csv
import gzip
from pathlib import Path

from openpyxl import load_workbook

KEEP = [
    "Scenario ID",
    "Height",
    "Weight",
    "Budget",
    "Primary Goal",
    "Intensity",
    "Foot & Calf Priority",
    "Space Constraint",
    "Primary Model",
    "Alternative Model 1",
    "Alternative Model 2",
    "Sales Priority (1–5)",
    "Recommendation Reason",
    "Main Trade-Off",
    "Do Not Recommend When",
]

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "data" / "sales"


def export_brand(brand: str, xlsx: Path) -> None:
    if not xlsx.is_file():
        raise SystemExit(f"missing workbook: {xlsx}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["All_Practical_Cases"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    idx = {h: i for i, h in enumerate(headers) if h}
    out = OUT_DIR / f"practical_cases_{brand}.csv.gz"
    with gzip.open(out, "wt", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=KEEP)
        writer.writeheader()
        n = 0
        for row in rows:
            n += 1
            writer.writerow({c: row[idx[c]] if c in idx else "" for c in KEEP})
    print(f"{brand}: {n} cases → {out} ({out.stat().st_size} bytes)")

    ws = wb["Active_Models"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    active_out = OUT_DIR / f"active_models_{brand}.csv"
    with active_out.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=["no", "name", "priority", "notes", "list_price"]
        )
        writer.writeheader()
        n = 0
        for row in rows:
            if not row or not row[1]:
                continue
            n += 1
            price = None
            for cell in row[2:6]:
                if isinstance(cell, (int, float)) and cell > 100:
                    price = cell
                    break
            pri = row[2]
            notes = row[3]
            writer.writerow(
                {
                    "no": row[0],
                    "name": row[1],
                    "priority": pri if isinstance(pri, (int, float)) and pri <= 10 else "",
                    "notes": notes if isinstance(notes, str) else ("" if notes is None else notes),
                    "list_price": price or "",
                }
            )
    print(f"{brand}: {n} active models → {active_out}")
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--osaki",
        type=Path,
        default=ROOT / "raw_data" / "sales" / "osaki_practical_cases.xlsx",
    )
    parser.add_argument(
        "--titan",
        type=Path,
        default=ROOT / "raw_data" / "sales" / "titan_practical_cases.xlsx",
    )
    args = parser.parse_args()
    export_brand("osaki", args.osaki)
    export_brand("titan", args.titan)


if __name__ == "__main__":
    main()
