#!/usr/bin/env python3
"""Export runtime fit specs from the Massage Chair specification workbook.

Writes ``data/sales/spec_index.json`` keyed by sales / case-book model names
(plus aliases) so recommend filtering does not depend on reason-text parsing.

Usage:
  python script/export_sales_spec_index.py
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any, Optional

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

from openpyxl import load_workbook  # noqa: E402

SPEC_XLSX = ROOT / "raw_data" / "Specification_Massage Chair.xlsx"
OUT_PATH = ROOT / "data" / "sales" / "spec_index.json"

# Case-book sales name → (spec Brand, spec Name) — shared with refill scripts.
MAP = {
    "Titan 3D Quantum": ("Titan", "3D Quantum"),
    "Osaki Oasis": ("Osaki", "Oasis"),
    "Titan Rejuv 4D": ("Titan", "Rejūv 4D"),
    "Osaki 4D Helix LE": ("Osaki Platinum", "Helix LE"),
    "Titan TP-Epic 4D": ("Titan", "TP Epic 4D"),
    "Osaki JP-Nexus 4D Made in Japan": ("Osaki Japan", "JP-Nexus 4D"),
    "Osaki 4D+3D Manhattan Duo": ("Osaki", "Manhattan Duo"),
    "Osaki OS-Trion Flex Duo 4D+3D": ("Osaki", "Trion Flex Duo"),
    "Osaki Platinum - Escape Duo 4D": ("Osaki Platinum", "4D Escape Duo"),
    "Osaki OS-Champ": ("Osaki", "Champ"),
    "Osaki Signature II": ("Osaki", "Signature II"),
    "Osaki Nova II 3D+": ("Osaki", "Nova II"),
    "OS-3D AI Vito": ("Osaki", "Vito"),
    "Titan Malibu Sync": ("Titan", "Malibu Sync"),
    "Ventura 3D": ("Osaki", "Ventura"),
    "Grande XL-Big and Tall": ("Titan", "Grande XL"),
    "Titan Pro 4D Astro": ("Titan", "Pro 4D Astro"),
    "Titan Pro Cura 4D": ("Titan", "Cura"),
    "Osaki Virtus Duo 4D": ("Osaki", "Virtus"),
    "Osaki aI 4D Yoga Flex": ("Osaki", "4D Yoga Flex"),
    "Osaki Pro 4D Epic LE": ("Osaki", "Epic LE"),
    "Osaki OS-Highpointe 4D": ("Osaki", "Highpointe"),
    "Osaki 5D+4D Kairos Duo": ("Osaki Platinum", "Kairos 5D+4D"),
    "Osaki 4D Bravo Duo Mech 4D+3D": ("Osaki", "Bravo Duo Mech"),
    "Osaki Duke XL 4D": ("Osaki", "Duke XL"),
    "Osaki 4D+3D Bravo Duo Flex": ("Osaki Platinum", "Bravo Duo Flex"),
    "Osaki 4D Maestro LE 2.0": ("Osaki", "Maestro LE 2.0"),
    "Osaki OP-4D Master": ("Osaki Platinum", "Master"),
    "Osaki OP-AI Xrest 4D": ("Osaki Platinum", "Xrest"),
    "Osaki AI Apex 5D+4D Duo": ("Osaki", "Apex Duo"),
    "Osaki OS-Pro 4D+3D DuoMax": ("Osaki", "DuoMax"),
    "Osaki OS-Pro 4D+3D DuoMax SE": ("Osaki Platinum", "4D DuoMax SE"),
    "Pinnacle 5D Duoflex AI": ("Osaki Platinum", "Pinnacle 5D DuoFlex AI"),
}

KAIROS_MAX_USER = 280.0

SPEC_OVERRIDES = {
    "Osaki OS-Highpointe 4D": {"door_asm": 36.5, "door_dis": None},
    "Osaki OS-Champ": {"max_user": 260},
}

# Extra lookup aliases (case / Shopify / customer phrasing).
EXTRA_ALIASES = {
    "Osaki aI 4D Yoga Flex": ["Osaki 4D Yoga Flex"],
    "Grande XL-Big and Tall": ["Titan Grande XL", "Grande XL"],
    "OS-3D AI Vito": ["Osaki OS-3D AI Vito", "Osaki Vito"],
    "Titan Rejuv 4D": ["Titan Rejūv 4D"],
    "Ventura 3D": ["Osaki Ventura 3D", "Osaki Ventura"],
}


def _num(v) -> Optional[float]:
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    match = re.search(r"(\d+(?:\.\d+)?)", str(v).replace(",", ""))
    return float(match.group(1)) if match else None


def main() -> None:
    if not SPEC_XLSX.is_file():
        raise SystemExit(f"missing spec workbook: {SPEC_XLSX}")

    wb = load_workbook(SPEC_XLSX, read_only=True, data_only=True)
    ws = wb["Massage Chair"]
    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        brand = row[0]
        name = row[1]
        if not name:
            continue
        by_key[(str(brand), str(name))] = {
            "brand": str(brand),
            "spec_name": str(name),
            "max_user_lb": _num(row[63]),  # col 64
            "wall_clearance_in": _num(row[40]),  # col 41
            "door_asm_in": _num(row[64]),  # col 65
            "door_dis_in": _num(row[65]),  # col 66
        }

    models: list[dict[str, Any]] = []
    missing: list[str] = []
    for sales_name, key in MAP.items():
        row = by_key.get(key)
        if row is None:
            missing.append(f"{sales_name} → {key}")
            continue
        entry = {
            "name": sales_name,
            "brand": row["brand"],
            "spec_name": row["spec_name"],
            "max_user_lb": row["max_user_lb"],
            "wall_clearance_in": row["wall_clearance_in"],
            "door_asm_in": row["door_asm_in"],
            "door_dis_in": row["door_dis_in"],
            "aliases": list(EXTRA_ALIASES.get(sales_name, [])),
        }
        if sales_name == "Osaki 5D+4D Kairos Duo":
            entry["max_user_lb"] = KAIROS_MAX_USER
        ov = SPEC_OVERRIDES.get(sales_name) or {}
        if "door_asm" in ov:
            entry["door_asm_in"] = ov["door_asm"]
        if "door_dis" in ov:
            entry["door_dis_in"] = ov["door_dis"]
        if "max_user" in ov:
            entry["max_user_lb"] = float(ov["max_user"])
        # Also index under the sheet display name.
        if row["spec_name"] and row["spec_name"] not in entry["aliases"]:
            entry["aliases"].append(row["spec_name"])
        branded = f"{row['brand']} {row['spec_name']}".strip()
        if branded and branded not in entry["aliases"] and branded != sales_name:
            entry["aliases"].append(branded)
        models.append(entry)

    if missing:
        raise SystemExit("missing MAP rows in spec sheet:\n  " + "\n  ".join(missing))

    # Also include every Active-ish sheet row not already covered, for specs Q&A.
    covered = {(m["brand"], m["spec_name"]) for m in models}
    for key, row in by_key.items():
        if key in covered:
            continue
        if row["door_asm_in"] is None and row["door_dis_in"] is None and row["max_user_lb"] is None:
            continue
        models.append(
            {
                "name": f"{row['brand']} {row['spec_name']}".strip(),
                "brand": row["brand"],
                "spec_name": row["spec_name"],
                "max_user_lb": row["max_user_lb"],
                "wall_clearance_in": row["wall_clearance_in"],
                "door_asm_in": row["door_asm_in"],
                "door_dis_in": row["door_dis_in"],
                "aliases": [row["spec_name"]],
            }
        )

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": str(SPEC_XLSX.name),
        "model_count": len(models),
        "models": models,
    }
    OUT_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    with_door = sum(1 for m in models if m.get("door_asm_in") or m.get("door_dis_in"))
    print(f"wrote {OUT_PATH} ({len(models)} models, {with_door} with doorway)")


if __name__ == "__main__":
    main()
