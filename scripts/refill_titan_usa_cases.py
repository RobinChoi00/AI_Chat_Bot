#!/usr/bin/env python3
"""Fill Titan & USA.xlsx All_Practical_Cases from priced Active_Models + spec sheet."""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

import refill_practical_cases as rc
from refill_practical_cases import (
    BUDGET_RANGE,
    Model,
    _box,
    _family,
    _foot_count,
    _fmt_in,
    _mech_rank,
    _num,
    _yes,
    do_not_text,
    reason_text,
    score_model,
    tradeoff_text,
)

SALES_XLSX = str(Path(__file__).resolve().parent.parent / "raw_data" / "sales" / "titan_practical_cases.xlsx")
SPEC_XLSX = rc.SPEC_XLSX

# Excel Active_Models name -> spec sheet (brand, name)
MAP = {
    "Titan eCabin 3D": ("Titan", "eCabin"),
    "OS-3D AI Vito": ("Osaki", "Vito"),
    "Titan Axiom LE": ("Titan", "Axiom LE"),
    "AmaMedic 3D Astoria": ("Amamedic", "3D Astoria"),
    "OS-3D Champ II": ("Osaki", "3D Champ 2"),
    "Osaki AI Monarch LE": ("Osaki", "Monarch LE"),
    "Titan Ignite Sync 3D+2D": ("Titan", "Ignite Sync"),
    "Grande XL-Big and Tall": ("Titan", "Grande XL"),
    "Osaki OS-3D Hamilton LE": ("Osaki", "Hamilton LE"),
    "Amamedic Haven 4D": ("Amamedic", "Pro 4D Haven"),
    "Titan Gemini": ("Titan", "Gemini"),
    "Titan Pro 4D Endor": ("Titan", "Pro 4D Endor"),
    "Titan Rejuv 4D": ("Titan", "Rejūv 4D"),
    "Osaki 4D Helix LE": ("Osaki Platinum", "Helix LE"),
    "Titan 4D Ion": ("Titan", "ion"),
    "Osaki Vibe 4D": ("Osaki", "Vibe 4D"),
    "Osaki OS-4D Achilles": ("Osaki", "4D Achilles"),
    "Osaki OP-4D Ultima": ("Osaki", "Ultima"),
    "Osaki Solo Flex 4D": ("Osaki", "Solo Flex"),
    "Osaki aI 4D Yoga Flex": ("Osaki", "4D Yoga Flex"),
    "Titan TP-Epic 4D": ("Titan", "TP Epic 4D"),
    "AmaMedic AI Revive 4D": ("Amamedic", "Revive"),
    "Osaki 4D Emperor II": ("Osaki", "Emperor II"),
    "Osaki Pro 4D Epic LE": ("Osaki", "Epic LE"),
    "Osaki OS-Highpointe 4D": ("Osaki", "Highpointe"),
    "Osaki 5D+4D Kairos Duo": ("Osaki Platinum", "Kairos 5D+4D"),
    "Osaki 4D Bravo Duo Mech 4D+3D": ("Osaki", "Bravo Duo Mech"),
    "Osaki Duke XL 4D": ("Osaki", "Duke XL"),
    "Osaki 4D+3D Bravo Dul Flex": ("Osaki Platinum", "Bravo Duo Flex"),
    "Osaki JP-Nexus 4D Made in Japan": ("Osaki Japan", "JP-Nexus 4D"),
    "Osaki 4D+3D Manhattan Duo": ("Osaki", "Manhattan Duo"),
    "Osaki 4D Maestro LE 2.0": ("Osaki", "Maestro LE 2.0"),
    "Osaki OP-4D Master": ("Osaki Platinum", "Master"),
    "Osaki OP-AI Xrest 4D": ("Osaki Platinum", "Xrest"),
    "Osaki OS-Trion Flex Duo 4D+3D": ("Osaki", "Trion Flex Duo"),
    "Osaki AI Apex 5D+4D Duo": ("Osaki", "Apex Duo"),
    "Osaki OS-Pro 4D+3D DuoMax": ("Osaki", "DuoMax"),
    "Osaki Platinum - Sapphire 4D+": ("Osaki Platinum", "4D Sapphire"),
    "Osaki OS-Pro 4D+3D DuoMax SE": ("Osaki Platinum", "4D DuoMax SE"),
    "Osaki Platinum - Escape Duo 4D": ("Osaki Platinum", "4D Escape Duo"),
    "Osaki Platinum - Vanguard Duo 4D": ("Osaki Platinum", "Vanguard Duo"),
    "Pinnacle 5D Duoflex AI": ("Osaki Platinum", "Pinnacle 5D DuoFlex AI"),
}

SPEC_OVERRIDES = {
    "Osaki OS-Highpointe 4D": {"door_asm": 36.5, "door_dis": None},
    "Osaki OS-4D Achilles": {"max_user": 260},  # official retailer spec; sheet blank
    "Osaki 5D+4D Kairos Duo": {"max_user": 280},  # sheet blank; retailer 280
}

REVIEW_MODELS = {
    "Osaki OS-4D Achilles",  # weight filled from retailer, not internal sheet
    "Osaki 5D+4D Kairos Duo",
    "Osaki Platinum - Sapphire 4D+",  # wall clearance blank on sheet
}

rc.XL_NAMES = {"Grande XL-Big and Tall", "Osaki Duke XL 4D"}
rc.SHORT_FRAME = {
    "Titan Axiom LE",
    "Titan Gemini",
    "Osaki Oasis",
}
rc.FAMILIES = [
    frozenset(
        {
            "Osaki 4D Bravo Duo Mech 4D+3D",
            "Osaki 4D+3D Bravo Dul Flex",
            "Osaki 4D+3D Bravo Duo Flex",
        }
    ),
    frozenset({"Osaki OS-Pro 4D+3D DuoMax", "Osaki OS-Pro 4D+3D DuoMax SE"}),
    frozenset(
        {"Osaki Platinum - Escape Duo 4D", "Osaki Platinum - Vanguard Duo 4D"}
    ),
]


def load_models() -> list[Model]:
    am = load_workbook(SALES_XLSX, data_only=True)["Active_Models"]
    targets = []
    for r in am.iter_rows(min_row=2, values_only=True):
        name = r[1]
        notes = r[3]
        if not name:
            continue
        if not isinstance(notes, (int, float)):
            continue
        targets.append(
            {
                "name": str(name).strip(),
                "focus": "Primary sales focus",
                "price": float(notes),
                "priority": 3,
            }
        )

    spec = load_workbook(SPEC_XLSX, data_only=True)["Massage Chair"]
    by_key = {}
    for r in range(5, spec.max_row + 1):
        brand, name = spec.cell(r, 1).value, spec.cell(r, 2).value
        if name:
            by_key[(str(brand), str(name))] = r

    models: list[Model] = []
    missing = []
    for t in targets:
        key = MAP.get(t["name"])
        if not key or key not in by_key:
            missing.append(t["name"])
            continue
        r = by_key[key]
        max_user = _num(spec.cell(r, 64).value)
        ov = SPEC_OVERRIDES.get(t["name"], {})
        if "max_user" in ov:
            max_user = float(ov["max_user"])
        if max_user is None:
            missing.append(t["name"] + " (no weight)")
            continue
        wall = spec.cell(r, 41).value
        wall_n = _num(wall)
        if str(wall).strip() in ("-", "None", ""):
            wall_n = None
        door_asm = _num(spec.cell(r, 65).value)
        door_dis = _num(spec.cell(r, 66).value)
        if "door_asm" in ov:
            door_asm = ov["door_asm"]
        if "door_dis" in ov:
            door_dis = ov["door_dis"]
        models.append(
            Model(
                name=t["name"],
                focus=t["focus"],
                price=t["price"],
                sales_priority=t["priority"],
                mech=str(spec.cell(r, 11).value or ""),
                track=str(spec.cell(r, 16).value or ""),
                track_len=_num(spec.cell(r, 18).value),
                air=_num(spec.cell(r, 19).value),
                auto=_num(spec.cell(r, 27).value),
                foot_raw=str(spec.cell(r, 33).value or ""),
                foot_n=_foot_count(spec.cell(r, 33).value),
                foot_ext=_num(spec.cell(r, 34).value),
                calf_knead=_yes(spec.cell(r, 36).value),
                calf_roller=_yes(spec.cell(r, 37).value),
                heat=str(spec.cell(r, 38).value or ""),
                zg=_num(spec.cell(r, 40).value),
                space=wall_n,
                ai=_yes(spec.cell(r, 49).value),
                seat_w=_num(spec.cell(r, 62).value),
                shoulder_w=_num(spec.cell(r, 61).value),
                max_user=max_user,
                door_asm=door_asm,
                door_dis=door_dis,
                nbox=int(_num(spec.cell(r, 67).value) or 1),
                pack=(
                    int(_num(spec.cell(r, 67).value) or 1),
                    _box(spec.cell(r, 69).value, spec.cell(r, 70).value, spec.cell(r, 71).value),
                    _box(spec.cell(r, 73).value, spec.cell(r, 74).value, spec.cell(r, 75).value),
                    _box(spec.cell(r, 77).value, spec.cell(r, 78).value, spec.cell(r, 79).value),
                ),
                spec_review=t["name"] in REVIEW_MODELS,
            )
        )
    if missing:
        raise SystemExit("unmapped models: " + ", ".join(missing))
    return models


def _is_weak_1d(m: Model) -> bool:
    mech = m.mech or ""
    return "1D" in mech and "3D" not in mech and "4D" not in mech and "5D" not in mech


def titan_hard_fail(m: Model, height, weight, space, budget_hi, foot, goal):
    why = rc.hard_fail(m, height, weight, space, budget_hi, foot or "Not Important", goal or "")
    if why:
        return why
    # Haven (23 in seat) is XL-wide even though it is not branded XL.
    if m.seat_w and m.seat_w >= 22.5:
        if height.startswith("Petite") and weight not in ("261–300 lb", "301+ lb"):
            return "wide_petite"
        if height.startswith("Average") and weight not in ("261–300 lb", "301+ lb"):
            return "wide_average"
        if weight in ("≤180 lb", "181–220 lb") and not height.startswith("Extra Tall"):
            return "wide_light"
    return None


def titan_score(m, height, weight, budget, goal, intensity, foot, space) -> float:
    s = score_model(m, height, weight, budget, goal, intensity, foot, space)
    if _is_weak_1d(m):
        # Dual 1D is not a gentle 3D substitute, and score_model treats "Dual" as full-body.
        s -= 8.0
        if goal == "Full-Body Relaxation":
            s -= 3.0
        if intensity == "Strong":
            s -= 1.5
    if intensity == "Gentle" and 2.8 <= _mech_rank(m.mech) <= 3.4:
        s += 2.5
    if height.startswith("Petite") and m.seat_w:
        s += (21.0 - m.seat_w) * 0.6
    # Epic LE / DuoMax capacity should not beat a 260–270 chair for lighter bands.
    if weight not in ("261–300 lb", "301+ lb") and m.max_user >= 297:
        s -= 2.5
    if m.spec_review:
        s -= 0.4
    return s


def titan_tradeoff(m, height, weight, intensity, space, budget) -> str:
    t = tradeoff_text(m, height, weight, intensity, space, budget)
    extras = []
    if m.door_asm and m.door_asm > 32:
        extras.append(
            f"assembled min door is {_fmt_in(m.door_asm)} in — do not promise a narrow doorway"
        )
    if m.spec_review:
        extras.append("confirm live retailer spec (weight and/or wall) before promising")
    if not extras:
        return t
    if t.startswith("No major spec trade-off"):
        return "; ".join(extras) + "."
    return t.rstrip(".") + "; " + "; ".join(extras) + "."


def pick_three(models, height, weight, budget, goal, intensity, foot, space):
    lo, hi = BUDGET_RANGE[budget]
    foot = foot or "Not Important"
    passing = [
        m
        for m in models
        if not titan_hard_fail(m, height, weight, space, hi, foot, goal)
    ]
    in_band = [m for m in passing if lo <= m.price <= hi]
    below = [m for m in passing if m.price < lo]
    pool = list(in_band) if in_band else list(below)
    if not pool:
        return []
    scored = [
        (titan_score(m, height, weight, budget, goal, intensity, foot, space), m)
        for m in pool
    ]
    # Cheaper in-band chair wins ties — do not upsell when specs are equal.
    scored.sort(key=lambda x: (-x[0], x[1].price, x[1].name))
    primary = scored[0][1]
    picks = [primary]
    used = {primary.name}
    used_fam = set(_family(primary.name))
    rest = [(s, m) for s, m in scored if m.name not in used]
    skipped = []
    for s, m in rest:
        if len(picks) >= 3:
            break
        fam = _family(m.name)
        if fam & used_fam:
            skipped.append(m)
            continue
        picks.append(m)
        used.add(m.name)
        used_fam |= set(fam)
    for m in skipped:
        if len(picks) >= 3:
            break
        if m.name not in used:
            picks.append(m)
    return picks


def preview(models):
    cases = [
        ("Petite (<5'4\")", "≤180 lb", "Under $3,000", "Neck & Shoulders", "Gentle", "Not Important", "No Space Constraint"),
        ("Petite (<5'4\")", "≤180 lb", "Under $3,000", "Foot & Calf", "Balanced", "Top Priority", "Small Room"),
        ("Average (5'4\"–5'11\")", "181–220 lb", "Under $3,000", "Foot & Calf", "Balanced", "Top Priority", "Narrow Doorway"),
        ("Average (5'4\"–5'11\")", "181–220 lb", "$3,000–$4,999", "Stretching & Mobility", "Balanced", "Not Important", "No Space Constraint"),
        ("Average (5'4\"–5'11\")", "221–260 lb", "$5,000–$6,999", "Neck & Shoulders", "Highly Adjustable", "Important", "Small Room"),
        ("Average (5'4\"–5'11\")", "221–260 lb", "$5,000–$6,999", "Foot & Calf", "Strong", "Top Priority", "Narrow Doorway"),
        ("Tall (6'0\"–6'2\")", "221–260 lb", "$7,000–$9,999", "Lower Back", "Strong", "Important", "No Space Constraint"),
        ("Extra Tall (6'3\"+)", "301+ lb", "Under $3,000", "Full-Body Relaxation", "Balanced", "Important", "No Space Constraint"),
        ("Extra Tall (6'3\"+)", "301+ lb", "$10,000+", "Full-Body Relaxation", "Strong", "Top Priority", "Narrow Doorway"),
        ("Petite (<5'4\")", "≤180 lb", "$3,000–$4,999", "Neck & Shoulders", "Gentle", "Not Important", "Narrow Doorway"),
        ("Average (5'4\"–5'11\")", "261–300 lb", "$3,000–$4,999", "Full-Body Relaxation", "Balanced", "Important", "No Space Constraint"),
        ("Tall (6'0\"–6'2\")", "181–220 lb", "Under $3,000", "Hips & Seat", "Balanced", "Important", "Small Room"),
        ("Average (5'4\"–5'11\")", "221–260 lb", "$3,000–$4,999", "Full-Body Relaxation", "Gentle", "Not Important", "Small Room"),
        ("Average (5'4\"–5'11\")", "221–260 lb", "$3,000–$4,999", "Neck & Shoulders", "Strong", "Important", "Narrow Doorway"),
        ("Average (5'4\"–5'11\")", "261–300 lb", "Under $3,000", "Full-Body Relaxation", "Balanced", "Important", "No Space Constraint"),
        ("Average (5'4\"–5'11\")", "261–300 lb", "$5,000–$6,999", "Lower Back", "Strong", "Important", "Small Room"),
        ("Extra Tall (6'3\"+)", "301+ lb", "$5,000–$6,999", "Lower Back", "Strong", "Top Priority", "Narrow Doorway"),
        ("Petite (<5'4\")", "≤180 lb", "Under $3,000", "Neck & Shoulders", "Gentle", "Not Important", "Small Room"),
        ("Petite (<5'4\")", "≤180 lb", "Under $3,000", "Neck & Shoulders", "Strong", "Not Important", "No Space Constraint"),
        ("Average (5'4\"–5'11\")", "181–220 lb", "$10,000+", "Full-Body Relaxation", "Strong", "Top Priority", "Small Room"),
        ("Tall (6'0\"–6'2\")", "221–260 lb", "$7,000–$9,999", "Foot & Calf", "Strong", "Top Priority", "Narrow Doorway"),
    ]
    print("\n===== PREVIEW =====")
    for c in cases:
        picks = pick_three(models, *c)
        names = [p.name for p in picks] if picks else ["NO VERIFIED MATCH"]
        print(f"{c[0][:12]:12} {c[1]:10} {c[2]:16} {c[3][:14]:14} {c[6]:16} -> {names}")


def write_sheet(models, preview_only=False):
    if preview_only:
        preview(models)
        return
    wb = load_workbook(SALES_XLSX)
    ws = wb["All_Practical_Cases"]
    n_empty = n_filled = n_review = 0
    primary_counts: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        height = ws.cell(r, 2).value
        weight = ws.cell(r, 3).value
        budget = ws.cell(r, 4).value
        goal = ws.cell(r, 5).value
        intensity = ws.cell(r, 6).value
        foot = ws.cell(r, 7).value
        space = ws.cell(r, 8).value
        picks = pick_three(models, height, weight, budget, goal, intensity, foot, space)

        def clear_pack(start):
            for c in range(start, start + 4):
                ws.cell(r, c).value = None

        if not picks:
            n_empty += 1
            ws.cell(r, 9).value = "NO VERIFIED MATCH"
            ws.cell(r, 10).value = None
            ws.cell(r, 11).value = None
            ws.cell(r, 12).value = None
            ws.cell(r, 13).value = (
                "No priced Titan/USA Active model in this budget has a verified "
                "weight capacity, assembled doorway, wall clearance, and foot/calf fit. Do not guess."
            )
            ws.cell(r, 14).value = "Leaving this case empty is more accurate than forcing an unfit model."
            ws.cell(r, 15).value = (
                "Do not recommend any listed model until a verified size/door/weight match exists in this budget."
            )
            ws.cell(r, 16).value = "Yes"
            for start in (17, 21, 25):
                clear_pack(start)
            continue

        n_filled += 1
        p = picks[0]
        primary_counts[p.name] = primary_counts.get(p.name, 0) + 1
        ws.cell(r, 9).value = p.name
        ws.cell(r, 10).value = picks[1].name if len(picks) > 1 else None
        ws.cell(r, 11).value = picks[2].name if len(picks) > 2 else None
        ws.cell(r, 12).value = None
        ws.cell(r, 13).value = reason_text(p, height, weight, budget, goal, intensity, foot or "Not Important", space)
        ws.cell(r, 14).value = titan_tradeoff(p, height, weight, intensity, space, budget)
        ws.cell(r, 15).value = do_not_text(p)
        review = any(x.spec_review for x in picks)
        ws.cell(r, 16).value = "Yes" if review else "No"
        if review:
            n_review += 1
        for i, start in enumerate((17, 21, 25)):
            if i < len(picks):
                nbox, b1, b2, b3 = picks[i].pack
                ws.cell(r, start).value = nbox
                ws.cell(r, start + 1).value = b1
                ws.cell(r, start + 2).value = b2
                ws.cell(r, start + 3).value = b3
            else:
                clear_pack(start)
        if r % 4000 == 0:
            print(f"  row {r} ...")
    wb.save(SALES_XLSX)
    print(f"filled={n_filled} empty={n_empty} spec_review={n_review}")
    print("primary distribution:")
    for k, v in sorted(primary_counts.items(), key=lambda kv: -kv[1]):
        print(f"  {v:5}  {k}")
    audit(ws)


NARROW_NEVER = {
    "Osaki OS-Highpointe 4D",
    "Osaki 5D+4D Kairos Duo",
    "Osaki Duke XL 4D",
    "Grande XL-Big and Tall",
    "Osaki AI Apex 5D+4D Duo",
    "Osaki JP-Nexus 4D Made in Japan",
    "Osaki OS-Trion Flex Duo 4D+3D",
    "Osaki Platinum - Sapphire 4D+",
}

LIGHT_MAX = {
    "Osaki Solo Flex 4D": 250,
    "Osaki 4D Emperor II": 250,
    "Titan Axiom LE": 240,
}


def audit(ws) -> None:
    print("\n===== AUDIT =====")
    issues = []
    narrow_hits = {n: 0 for n in NARROW_NEVER}
    light_on_260 = {n: 0 for n in LIGHT_MAX}
    cap270_on_261 = 0
    for r in range(2, ws.max_row + 1):
        weight = ws.cell(r, 3).value
        space = ws.cell(r, 8).value
        names = [ws.cell(r, c).value for c in (9, 10, 11)]
        for n in names:
            if not n or n == "NO VERIFIED MATCH":
                continue
            if space == "Narrow Doorway" and n in narrow_hits:
                narrow_hits[n] += 1
            if weight == "221–260 lb" and n in light_on_260:
                light_on_260[n] += 1
            if weight == "261–300 lb" and n in {
                "Titan eCabin 3D",
                "Osaki OS-3D Hamilton LE",
                "Amamedic Haven 4D",
                "Titan 4D Ion",
                "Osaki Vibe 4D",
                "Osaki OP-4D Ultima",
                "AmaMedic AI Revive 4D",
                "Osaki OS-Highpointe 4D",
                "Osaki 4D Bravo Duo Mech 4D+3D",
                "Osaki 4D+3D Bravo Dul Flex",
                "Osaki OS-Trion Flex Duo 4D+3D",
            }:
                cap270_on_261 += 1
    for n, c in narrow_hits.items():
        status = "OK" if c == 0 else "FAIL"
        print(f"  {status} Narrow hits for {n}: {c}")
        if c:
            issues.append(n)
    for n, c in light_on_260.items():
        status = "OK" if c == 0 else "FAIL"
        print(f"  {status} 221–260 hits for {n}: {c}")
        if c:
            issues.append(n)
    print(f"  {'OK' if cap270_on_261 == 0 else 'FAIL'} 270-cap chairs on 261–300: {cap270_on_261}")
    if issues or cap270_on_261:
        print("AUDIT HAD FAILURES")
    else:
        print("AUDIT PASSED")


if __name__ == "__main__":
    models = load_models()
    print(f"loaded {len(models)} priced models")
    for m in sorted(models, key=lambda x: x.price):
        print(
            f"  ${m.price:7.0f} cap={m.max_user:5.0f} door={m.door_asm} wall={m.space} "
            f"foot={m.foot_n} calf={int(m.calf_strong)}  {m.name}"
        )
    preview_only = "--preview" in sys.argv
    write_sheet(models, preview_only=preview_only)
