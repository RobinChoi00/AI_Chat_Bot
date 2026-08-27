#!/usr/bin/env python3
"""Rebuild All_Practical_Cases recommendations from verified spec-sheet data."""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Optional

from openpyxl import load_workbook

from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
SPEC_XLSX = str(_ROOT / "raw_data" / "Specification_Massage Chair.xlsx")
SALES_XLSX = str(_ROOT / "raw_data" / "sales" / "osaki_practical_cases.xlsx")

MAP = {
    "Titan 3D Quantum": ("Titan", "3D Quantum"),
    "Osaki Oasis": ("Osaki", "Oasis"),
    "Titan Rejuv 4D": ("Titan", "Rejūv 4D"),
    "Osaki 4D Helix LE": ("Osaki Platinum", "Helix LE"),
    "Titan TP-Epic 4D": ("Titan", "TP Epic 4D"),
    "Osaki JP-Nexus 4D Made in Japan": ("Osaki Japan", "JP-Nexus 4D"),
    "Osaki 4D+3D Manhattan Duo": ("Osaki", "Manhattan Duo"),
    "Osaki OS-Trion Flex Duo 4D+3D": ("Osaki", "Trion Flex Duo"),
    "Osaki Platinum - Escape Duo 4D": ("Osaki Platinum", "4D Escape Duo"),
    "Osaki OS-Champ": ("Osaki", "Champ"),
    "Osaki Signature II": ("Osaki", "Signature II"),
    "Osaki Nova II 3D+": ("Osaki", "Nova II"),
    "OS-3D AI Vito": ("Osaki", "Vito"),
    "Titan Malibu Sync": ("Titan", "Malibu Sync"),
    "Ventura 3D": ("Osaki", "Ventura"),
    "Grande XL-Big and Tall": ("Titan", "Grande XL"),
    "Titan Pro 4D Astro": ("Titan", "Pro 4D Astro"),
    "Titan Pro Cura 4D": ("Titan", "Cura"),
    "Osaki Virtus Duo 4D": ("Osaki", "Virtus"),
    "Osaki aI 4D Yoga Flex": ("Osaki", "4D Yoga Flex"),
    "Osaki Pro 4D Epic LE": ("Osaki", "Epic LE"),
    "Osaki OS-Highpointe 4D": ("Osaki", "Highpointe"),
    "Osaki 5D+4D Kairos Duo": ("Osaki Platinum", "Kairos 5D+4D"),
    "Osaki 4D Bravo Duo Mech 4D+3D": ("Osaki", "Bravo Duo Mech"),
    "Osaki Duke XL 4D": ("Osaki", "Duke XL"),
    "Osaki 4D+3D Bravo Duo Flex": ("Osaki Platinum", "Bravo Duo Flex"),
    "Osaki 4D Maestro LE 2.0": ("Osaki", "Maestro LE 2.0"),
    "Osaki OP-4D Master": ("Osaki Platinum", "Master"),
    "Osaki OP-AI Xrest 4D": ("Osaki Platinum", "Xrest"),
    "Osaki AI Apex 5D+4D Duo": ("Osaki", "Apex Duo"),
    "Osaki OS-Pro 4D+3D DuoMax": ("Osaki", "DuoMax"),
    "Osaki OS-Pro 4D+3D DuoMax SE": ("Osaki Platinum", "4D DuoMax SE"),
    "Pinnacle 5D Duoflex AI": ("Osaki Platinum", "Pinnacle 5D DuoFlex AI"),
}

# Retailer-verified; missing from internal spec sheet.
KAIROS_MAX_USER = 280

# Sales-confirmed overrides when the spec sheet disagrees with the live doorway number.
SPEC_OVERRIDES = {
    "Osaki OS-Highpointe 4D": {
        "door_asm": 36.5,
        "door_dis": None,
    },
    # Official OTAS spec table lists 260 lb; sheet had 265. Use the lower number.
    "Osaki OS-Champ": {
        "max_user": 260,
    },
}

BUDGET_RANGE = {
    "Under $3,000": (0, 2999),
    "$3,000–$4,999": (3000, 4999),
    "$5,000–$6,999": (5000, 6999),
    "$7,000–$9,999": (7000, 9999),
    "$10,000+": (10000, 10**9),
}

# Required capacity = the top of the scenario band. A 270 lb chair is not a 261–300 lb chair.
WEIGHT_HARD = {
    "≤180 lb": 180,
    "181–220 lb": 220,
    "221–260 lb": 260,
    "261–300 lb": 297,  # Epic LE 297 is the lightest verified cover; 270–280 chairs do not cover this band
    "301+ lb": 330,
}

WEIGHT_IDEAL = {
    "≤180 lb": 180,
    "181–220 lb": 220,
    "221–260 lb": 260,
    "261–300 lb": 300,
    "301+ lb": 375,
}

NARROW_MAX_IN = 32.0  # effective assembled or disassembled
SMALL_ROOM_FAIL_IN = 6.0  # wall clearance > 6 in is not a small-room chair
XL_NAMES = {"Grande XL-Big and Tall", "Osaki Duke XL 4D"}


def _num(v) -> Optional[float]:
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    return float(m.group(1)) if m else None


def _foot_count(text) -> int:
    t = str(text or "").lower()
    if "tri" in t:
        return 3
    if t.startswith("4") or "4 roller" in t or "4 spinning" in t:
        return 4
    if t.startswith("3") or "3 roller" in t or "3 spinning" in t:
        return 3
    if t.startswith("2") or "2 roller" in t or "2 spinning" in t or "2 plate" in t:
        return 2
    if "acupuncture" in t or "ridge" in t:
        return 2
    if t.startswith("1") or "1 roller" in t or "1 spinning" in t:
        return 1
    return 1


def _yes(v) -> bool:
    s = str(v or "").strip().lower()
    return s not in ("", "-", "none", "no", "n", "0")


def _mech_rank(mech: str) -> float:
    m = str(mech or "")
    if "5D" in m:
        return 5.0
    if "Dual" in m and "4D" in m:
        return 4.6
    if "4D" in m:
        return 4.0
    if "3D+" in m:
        return 3.4
    if "3D" in m:
        return 3.0
    if "2D" in m:
        return 2.0
    if "1D" in m:
        return 2.0
    return 3.0


def _is_dual(mech: str) -> bool:
    return "dual" in str(mech or "").lower()


def _fmt_in(v: Optional[float]) -> str:
    if v is None:
        return ""
    if abs(v - round(v)) < 1e-6:
        return str(int(round(v)))
    return f"{v:.1f}".rstrip("0").rstrip(".")


def _box(w, l, h) -> Optional[str]:
    wn, ln, hn = _num(w), _num(l), _num(h)
    if wn is None or ln is None or hn is None:
        return None
    return f"{_fmt_in(wn)} × {_fmt_in(ln)} × {_fmt_in(hn)} in"


@dataclass
class Model:
    name: str
    focus: str  # Primary / Secondary
    price: float
    sales_priority: int
    mech: str
    track: str
    track_len: Optional[float]
    air: Optional[float]
    auto: Optional[float]
    foot_raw: str
    foot_n: int
    foot_ext: Optional[float]
    calf_knead: bool
    calf_roller: bool
    heat: str
    zg: Optional[float]
    space: Optional[float]
    ai: bool
    seat_w: Optional[float]
    shoulder_w: Optional[float]
    max_user: float
    door_asm: Optional[float]
    door_dis: Optional[float]
    nbox: Optional[int]
    pack: tuple
    spec_review: bool = False

    @property
    def is_primary(self) -> bool:
        return self.focus == "Primary sales focus"

    @property
    def is_xl(self) -> bool:
        return self.name in XL_NAMES

    @property
    def is_sl_flex(self) -> bool:
        return "flex" in str(self.track or "").lower()

    @property
    def is_straight(self) -> bool:
        return "straight" in str(self.track or "").lower()

    @property
    def door_effective(self) -> Optional[float]:
        """Assembled doorway only.

        Disassembled clearance is a delivery-tech option, not a reason to
        recommend the chair for a Narrow Doorway guest. Highpointe was the
        example: sheet listed 30 in disassembled, live min door is 36.5 in.
        """
        return self.door_asm or self.door_dis

    @property
    def door_listed(self) -> Optional[float]:
        return self.door_asm or self.door_dis

    @property
    def calf_strong(self) -> bool:
        return self.calf_knead or self.calf_roller

    @property
    def stretch_chair(self) -> bool:
        return "yoga" in self.name.lower()


def load_models() -> list[Model]:
    am = load_workbook(SALES_XLSX, data_only=True)["Active_Models"]
    targets = []
    for r in am.iter_rows(min_row=2, values_only=True):
        if r[3] in ("Primary sales focus", "Secondary sales focus") and r[1]:
            targets.append(
                {
                    "name": str(r[1]).strip(),
                    "focus": r[3],
                    "price": float(r[4] or 0),
                    "priority": int(r[2] or 3),
                }
            )

    spec = load_workbook(SPEC_XLSX, data_only=True)["Massage Chair"]
    by_key = {}
    for r in range(5, spec.max_row + 1):
        brand, name = spec.cell(r, 1).value, spec.cell(r, 2).value
        if name:
            by_key[(str(brand), str(name))] = r

    models: list[Model] = []
    for t in targets:
        key = MAP[t["name"]]
        r = by_key[key]
        max_user = _num(spec.cell(r, 64).value)
        spec_review = False
        if t["name"] == "Osaki 5D+4D Kairos Duo":
            max_user = float(KAIROS_MAX_USER)
            spec_review = True
        if max_user is None:
            raise SystemExit(f"missing max_user: {t['name']}")
        models.append(
            Model(
                name=t["name"],
                focus=t["focus"],
                price=t["price"],
                sales_priority=t["priority"],
                mech=str(spec.cell(r, 11).value or ""),
                track=str(spec.cell(r, 16).value or ""),
                track_len=_num(spec.cell(r, 18).value),
                air=_num(spec.cell(r, 19).value),
                auto=_num(spec.cell(r, 27).value),
                foot_raw=str(spec.cell(r, 33).value or ""),
                foot_n=_foot_count(spec.cell(r, 33).value),
                foot_ext=_num(spec.cell(r, 34).value),
                calf_knead=_yes(spec.cell(r, 36).value),
                calf_roller=_yes(spec.cell(r, 37).value),
                heat=str(spec.cell(r, 38).value or ""),
                zg=_num(spec.cell(r, 40).value),
                space=_num(spec.cell(r, 41).value),
                ai=_yes(spec.cell(r, 49).value),
                seat_w=_num(spec.cell(r, 62).value),
                shoulder_w=_num(spec.cell(r, 61).value),
                max_user=max_user,
                door_asm=_num(spec.cell(r, 65).value),
                door_dis=_num(spec.cell(r, 66).value),
                nbox=int(_num(spec.cell(r, 67).value) or 1),
                pack=(
                    int(_num(spec.cell(r, 67).value) or 1),
                    _box(spec.cell(r, 69).value, spec.cell(r, 70).value, spec.cell(r, 71).value),
                    _box(spec.cell(r, 73).value, spec.cell(r, 74).value, spec.cell(r, 75).value),
                    _box(spec.cell(r, 77).value, spec.cell(r, 78).value, spec.cell(r, 79).value),
                ),
                spec_review=spec_review,
            )
        )
        ov = SPEC_OVERRIDES.get(t["name"])
        if ov:
            m = models[-1]
            if "door_asm" in ov:
                m.door_asm = ov["door_asm"]
            if "door_dis" in ov:
                m.door_dis = ov["door_dis"]
            if "max_user" in ov:
                m.max_user = float(ov["max_user"])
    return models


SHORT_FRAME = {
    "Osaki OS-Champ",
    "Osaki Signature II",
    "Osaki Oasis",
}

FAMILIES = [
    frozenset({"Osaki 4D Bravo Duo Mech 4D+3D", "Osaki 4D+3D Bravo Duo Flex"}),
    frozenset({"Osaki OS-Pro 4D+3D DuoMax", "Osaki OS-Pro 4D+3D DuoMax SE"}),
]


def _family(name: str) -> frozenset:
    for fam in FAMILIES:
        if name in fam:
            return fam
    return frozenset({name})


def _foot_ok(m: Model, foot: str, goal: str) -> bool:
    """Spec-literal foot/calf gate."""
    need_top = foot == "Top Priority" or goal == "Foot & Calf"
    need_some = foot == "Important"
    if need_top:
        return m.foot_n >= 3 or (m.foot_n >= 2 and m.calf_strong)
    if need_some:
        return m.foot_n >= 2 or m.calf_strong
    return True


def hard_fail(
    m: Model,
    height: str,
    weight: str,
    space: str,
    budget_hi: float,
    foot: str = "Not Important",
    goal: str = "",
) -> Optional[str]:
    if m.price > budget_hi:
        return "over_budget"
    if m.max_user < WEIGHT_HARD[weight]:
        return "weight"
    if not _foot_ok(m, foot, goal):
        return "foot"
    if m.is_xl and height.startswith("Petite") and weight not in ("261–300 lb", "301+ lb"):
        return "xl_petite"
    if m.is_xl and height.startswith("Average") and weight not in ("261–300 lb", "301+ lb"):
        return "xl_average"
    if m.is_xl and weight in ("≤180 lb", "181–220 lb") and not height.startswith("Extra Tall"):
        return "xl_light"
    if m.is_straight and height.startswith("Extra Tall"):
        return "straight_tall"
    if height.startswith("Extra Tall") and m.name in SHORT_FRAME:
        return "too_short"
    if space == "Narrow Doorway":
        if m.door_effective is None or m.door_effective > NARROW_MAX_IN:
            return "door"
    if space == "Small Room":
        if m.space is None or m.space > SMALL_ROOM_FAIL_IN:
            return "wall"
    return None


def score_model(
    m: Model,
    height: str,
    weight: str,
    budget: str,
    goal: str,
    intensity: str,
    foot: str,
    space: str,
) -> float:
    """Rank only among chairs that already passed spec gates."""
    lo, hi = BUDGET_RANGE[budget]
    mr = _mech_rank(m.mech)
    s = 0.0
    if lo <= m.price <= hi:
        s += 20.0

    need = WEIGHT_HARD[weight]
    s += min(max(m.max_user - need, 0.0), 80.0) / 40.0

    if space == "Small Room" and m.space is not None:
        s += (SMALL_ROOM_FAIL_IN - m.space) * 2.0
    if space == "Narrow Doorway" and m.door_effective is not None:
        if m.door_asm and m.door_asm <= NARROW_MAX_IN:
            s += 3.0
        s += (NARROW_MAX_IN - m.door_effective) * 0.25

    if foot == "Top Priority" or goal == "Foot & Calf":
        s += m.foot_n * 2.0
        if m.calf_strong:
            s += 4.0
        if m.foot_ext:
            s += min(m.foot_ext, 10.0) * 0.15
    elif foot == "Important":
        s += m.foot_n * 0.6
        if m.calf_strong:
            s += 1.5

    if intensity == "Gentle":
        s += (4.2 - mr) * 2.5
    elif intensity == "Balanced":
        s += 2.0 if 3.0 <= mr <= 4.2 else 0.0
    elif intensity == "Strong":
        s += (mr - 2.8) * 2.0
        if _is_dual(m.mech):
            s += 1.5
    else:
        if m.ai:
            s += 2.0
        s += max(mr - 3.0, 0.0) * 1.5

    if goal == "Stretching & Mobility" and m.stretch_chair:
        s += 8.0
    elif goal == "Neck & Shoulders" and intensity != "Gentle" and "5D" in str(m.mech):
        s += 3.0
    elif goal in ("Lower Back", "Hips & Seat", "Upper Back"):
        if m.is_sl_flex:
            s += 3.0
        if m.track_len and m.track_len >= 54:
            s += 3.0
    elif goal == "Arms & Hands" and m.air:
        s += min(m.air, 60.0) / 20.0
    elif goal == "Full-Body Relaxation" and _is_dual(m.mech):
        s += 3.0

    if height.startswith("Extra Tall") and m.is_xl and weight in ("261–300 lb", "301+ lb"):
        s += 6.0
    if height.startswith("Petite") and m.is_xl:
        s -= 4.0
    if m.name == "Osaki Oasis":
        s -= 4.0

    s += (6 - min(m.sales_priority, 5)) * 0.1
    return s


def reason_text(m: Model, height, weight, budget, goal, intensity, foot, space) -> str:
    bits = [
        f"{m.mech} / {m.track}",
        f"{int(m.max_user)} lb capacity",
    ]
    if m.door_effective:
        bits.append(f"{_fmt_in(m.door_effective)} in min doorway")
    if m.space is not None:
        bits.append(f"{_fmt_in(m.space)} in wall clearance")
    foot_bit = f"{m.foot_n}-roller foot"
    if m.calf_strong:
        foot_bit += " + calf knead/roller"
    bits.append(foot_bit)
    why = []
    if goal == "Stretching & Mobility":
        why.append(
            "the dedicated Yoga Flex stretch frame"
            if m.stretch_chair
            else "stretch / zero-gravity range for mobility"
        )
    elif goal == "Foot & Calf":
        why.append("the strongest verified foot/calf package in this budget")
    elif goal == "Neck & Shoulders":
        why.append("4D/5D roller depth for neck and shoulders" if _mech_rank(m.mech) >= 4 else "neck/shoulder coverage in this budget")
    elif goal in ("Upper Back", "Lower Back"):
        if m.is_sl_flex:
            why.append("SL-Flex coverage through the glutes and lumbar")
        elif m.track_len and m.track_len >= 54:
            why.append("a long verified roller track for a taller spine")
        else:
            why.append(f"{goal.lower()} coverage on an SL-Track")
    elif goal == "Hips & Seat":
        why.append("SL-Flex / glute-path coverage" if m.is_sl_flex else "seat/hip airbag and roller coverage")
    elif goal == "Arms & Hands":
        why.append(f"{int(m.air or 0)} air cells for arm/hand compression")
    elif goal == "Full-Body Relaxation":
        why.append("dual-mechanism full-body coverage" if _is_dual(m.mech) else "full-body programs and air coverage")
    elif m.is_xl and (height.startswith("Extra Tall") or weight in ("261–300 lb", "301+ lb")):
        why.append("true big-and-tall seat/weight coverage")
    elif space == "Small Room" and m.space is not None and m.space <= 4:
        why.append("tight wall-clearance design")
    elif space == "Narrow Doorway" and m.door_effective and m.door_effective <= 30:
        why.append("a verified narrow delivery path")
    elif intensity == "Gentle" and _mech_rank(m.mech) <= 3.1:
        why.append("a milder 2D/3D roller that matches a gentle preference")
    elif intensity in ("Strong", "Highly Adjustable") and _mech_rank(m.mech) >= 4:
        why.append("adjustable 4D/dual intensity")
    if not why:
        why.append(f"the best verified spec match for {goal.lower()} at this budget")
    return (
        f"Lead with this model in {budget}: {'; '.join(bits)}. "
        f"Fitter note: {why[0]} for a {height.split()[0].lower()} / {weight} guest who wants {intensity.lower()} intensity."
    )


def tradeoff_text(m: Model, height, weight, intensity, space, budget) -> str:
    notes = []
    lo, hi = BUDGET_RANGE[budget]
    if m.price < lo:
        if m.stretch_chair:
            notes.append(
                f"Yoga Flex is ${int(m.price):,} — still the stretch chair; keep an in-band model as the upsell"
            )
        elif m.is_xl or m.max_user >= 330:
            notes.append(
                f"${int(m.price):,} is below this budget because it is the only verified size/weight fit"
            )
        else:
            notes.append(
                f"priced at ${int(m.price):,} which is below this budget band — it is the better verified fit"
            )
    if m.price and weight in ("261–300 lb", "301+ lb") and m.max_user < WEIGHT_IDEAL[weight]:
        notes.append(f"capacity is {int(m.max_user)} lb — confirm body weight before promising a fit")
    if m.is_xl and height.startswith("Average") and weight in ("≤180 lb", "181–220 lb"):
        notes.append("XL frame is oversized for a lighter average-height user")
    if m.is_xl and height.startswith("Petite"):
        notes.append("XL frame is a size compromise for a petite user; recommended only because of the weight need")
    if not m.is_xl and height.startswith("Extra Tall"):
        notes.append("not a dedicated XL frame — have the customer test recline/footrest length")
    if intensity == "Gentle" and _mech_rank(m.mech) >= 4.5:
        notes.append("dual 4D/5D can feel too aggressive if they truly want gentle")
    if intensity == "Strong" and _mech_rank(m.mech) <= 3.1:
        notes.append("2D/3D will not feel as deep as a 4D chair")
    if space == "Small Room" and m.space and m.space >= 6:
        notes.append(f"needs about {_fmt_in(m.space)} in from the wall")
    if space == "Narrow Doorway" and m.door_asm and m.door_dis and m.door_asm > 32:
        notes.append(
            f"assembled doorway is {_fmt_in(m.door_asm)} in; use the {_fmt_in(m.door_dis)} in disassembled path"
        )
    if not notes:
        return "No major spec trade-off vs other models that pass the verified height/weight/door filters."
    return "; ".join(notes) + "."


def do_not_text(m: Model) -> str:
    door = m.door_effective
    extra = ""
    if m.door_asm:
        extra = f" ({_fmt_in(m.door_asm)} in assembled"
        if m.door_dis:
            extra += f"; {_fmt_in(m.door_dis)} in only if fully disassembled — do not promise this for a narrow doorway"
        extra += ")"
    bits = [f"user exceeds {int(m.max_user)} lb"]
    if m.is_xl:
        bits.append("the user is petite / under ~5'4\" (XL seat and roller path will not land well)")
    if m.is_straight:
        bits.append("the user is extra-tall and needs glute/hamstring roller coverage")
    if door:
        bits.append(f"the delivery path is under {_fmt_in(door)} in{extra}")
    if m.space is not None:
        bits.append(f"the room cannot spare ~{_fmt_in(m.space)} in of wall clearance")
    if m.foot_n <= 1 and not m.calf_strong:
        bits.append("the guest needs a real foot/calf package (this model is 1-roller, no calf knead/roller)")
    return "Do not recommend if " + "; or ".join(bits) + "."


def pick_three(models: list[Model], height, weight, budget, goal, intensity, foot, space):
    lo, hi = BUDGET_RANGE[budget]
    passing = [
        m
        for m in models
        if not hard_fail(m, height, weight, space, hi, foot, goal)
    ]
    in_band = [m for m in passing if lo <= m.price <= hi]
    below = [m for m in passing if m.price < lo]

    # Stay in the quoted band whenever a spec-legal chair exists there.
    # Cheaper chairs are used only when the band has zero verified fits
    # (example: 301+ lb under $5k → Grande XL is the only 330+ chair).
    pool = list(in_band) if in_band else list(below)

    if not pool:
        return []

    scored = [
        (score_model(m, height, weight, budget, goal, intensity, foot, space), m) for m in pool
    ]
    scored.sort(key=lambda x: (-x[0], x[1].sales_priority, -x[1].price))
    primary = scored[0][1]

    picks = [primary]
    used = {primary.name}
    used_fam = set(_family(primary.name))
    rest = [(s, m) for s, m in scored if m.name not in used]

    if primary.price < lo:
        ib = [(s, m) for s, m in rest if lo <= m.price <= hi]
        if ib:
            picks.append(ib[0][1])
            used.add(ib[0][1].name)
            used_fam |= set(_family(ib[0][1].name))
            rest = [(s, m) for s, m in rest if m.name not in used]

    skipped = []
    for s, m in rest:
        if len(picks) >= 3:
            break
        fam = _family(m.name)
        if fam & used_fam:
            skipped.append(m)
            continue
        picks.append(m)
        used.add(m.name)
        used_fam |= set(fam)

    for m in skipped:
        if len(picks) >= 3:
            break
        if m.name not in used:
            picks.append(m)
            used.add(m.name)
    return picks


def preview(models: list[Model]) -> None:
    cases = [
        ("Petite (<5'4\")", "≤180 lb", "Under $3,000", "Neck & Shoulders", "Gentle", "Not Important", "No Space Constraint"),
        ("Petite (<5'4\")", "≤180 lb", "Under $3,000", "Stretching & Mobility", "Balanced", "Important", "Small Room"),
        ("Petite (<5'4\")", "≤180 lb", "$3,000–$4,999", "Stretching & Mobility", "Balanced", "Not Important", "No Space Constraint"),
        ("Average (5'4\"–5'11\")", "181–220 lb", "Under $3,000", "Foot & Calf", "Balanced", "Top Priority", "Narrow Doorway"),
        ("Average (5'4\"–5'11\")", "181–220 lb", "$3,000–$4,999", "Full-Body Relaxation", "Strong", "Important", "No Space Constraint"),
        ("Average (5'4\"–5'11\")", "221–260 lb", "$5,000–$6,999", "Neck & Shoulders", "Highly Adjustable", "Important", "Small Room"),
        ("Average (5'4\"–5'11\")", "221–260 lb", "$5,000–$6,999", "Foot & Calf", "Strong", "Top Priority", "Narrow Doorway"),
        ("Tall (6'0\"–6'2\")", "221–260 lb", "$7,000–$9,999", "Lower Back", "Strong", "Important", "No Space Constraint"),
        ("Tall (6'0\"–6'2\")", "261–300 lb", "$3,000–$4,999", "Full-Body Relaxation", "Balanced", "Important", "No Space Constraint"),
        ("Extra Tall (6'3\"+)", "301+ lb", "Under $3,000", "Full-Body Relaxation", "Balanced", "Important", "No Space Constraint"),
        ("Extra Tall (6'3\"+)", "301+ lb", "$5,000–$6,999", "Lower Back", "Strong", "Top Priority", "No Space Constraint"),
        ("Extra Tall (6'3\"+)", "301+ lb", "$10,000+", "Full-Body Relaxation", "Strong", "Top Priority", "No Space Constraint"),
        ("Extra Tall (6'3\"+)", "301+ lb", "$10,000+", "Full-Body Relaxation", "Strong", "Top Priority", "Narrow Doorway"),
        ("Petite (<5'4\")", "≤180 lb", "$10,000+", "Neck & Shoulders", "Gentle", "Not Important", "Small Room"),
        ("Average (5'4\"–5'11\")", "181–220 lb", "$7,000–$9,999", "Stretching & Mobility", "Highly Adjustable", "Not Important", "Small Room"),
        ("Average (5'4\"–5'11\")", "261–300 lb", "Under $3,000", "Upper Back", "Strong", "Important", "Narrow Doorway"),
        ("Tall (6'0\"–6'2\")", "181–220 lb", "$3,000–$4,999", "Hips & Seat", "Balanced", "Important", "Small Room"),
        ("Petite (<5'4\")", "221–260 lb", "$5,000–$6,999", "Arms & Hands", "Gentle", "Not Important", "No Space Constraint"),
    ]
    print("\n===== PREVIEW =====")
    for c in cases:
        picks = pick_three(models, *c)
        names = [p.name for p in picks] if picks else ["NO VERIFIED MATCH"]
        print(f"{c[0][:12]:12} {c[1]:10} {c[2]:16} {c[3][:16]:16} {c[4][:8]:8} {c[5][:8]:8} {c[6][:10]:10} -> {names}")


def write_sheet(models: list[Model], preview_only: bool = False) -> None:
    if preview_only:
        preview(models)
        return

    wb = load_workbook(SALES_XLSX)
    ws = wb["All_Practical_Cases"]
    n_empty = 0
    n_filled = 0
    n_review = 0
    primary_counts: dict[str, int] = {}

    for r in range(2, ws.max_row + 1):
        height = ws.cell(r, 2).value
        weight = ws.cell(r, 3).value
        budget = ws.cell(r, 4).value
        goal = ws.cell(r, 5).value
        intensity = ws.cell(r, 6).value
        foot = ws.cell(r, 7).value
        space = ws.cell(r, 8).value
        picks = pick_three(models, height, weight, budget, goal, intensity, foot, space)

        def clear_pack(start_col: int):
            for c in range(start_col, start_col + 4):
                ws.cell(r, c).value = None

        if not picks:
            n_empty += 1
            ws.cell(r, 9).value = "NO VERIFIED MATCH"
            ws.cell(r, 10).value = None
            ws.cell(r, 11).value = None
            ws.cell(r, 12).value = None
            ws.cell(r, 13).value = (
                "No Primary/Secondary model in this budget has a verified weight capacity, "
                "doorway, and size fit for this combination. Do not guess a chair."
            )
            ws.cell(r, 14).value = "Leaving this case empty is more accurate than forcing an unfit model."
            ws.cell(r, 15).value = (
                "Do not recommend any listed model until a verified XL/capacity/doorway match exists in this budget."
            )
            ws.cell(r, 16).value = "Yes"
            for start in (17, 21, 25):
                clear_pack(start)
            continue

        n_filled += 1
        p = picks[0]
        primary_counts[p.name] = primary_counts.get(p.name, 0) + 1
        ws.cell(r, 9).value = p.name
        ws.cell(r, 10).value = picks[1].name if len(picks) > 1 else None
        ws.cell(r, 11).value = picks[2].name if len(picks) > 2 else None
        ws.cell(r, 12).value = p.sales_priority
        ws.cell(r, 13).value = reason_text(p, height, weight, budget, goal, intensity, foot, space)
        ws.cell(r, 14).value = tradeoff_text(p, height, weight, intensity, space, budget)
        ws.cell(r, 15).value = do_not_text(p)
        review = any(x.spec_review for x in picks)
        ws.cell(r, 16).value = "Yes" if review else "No"
        if review:
            n_review += 1

        pack_starts = (17, 21, 25)
        for i, start in enumerate(pack_starts):
            if i < len(picks):
                nbox, b1, b2, b3 = picks[i].pack
                ws.cell(r, start).value = nbox
                ws.cell(r, start + 1).value = b1
                ws.cell(r, start + 2).value = b2
                ws.cell(r, start + 3).value = b3
            else:
                clear_pack(start)

        if r % 4000 == 0:
            print(f"  row {r} ...")

    wb.save(SALES_XLSX)
    print(f"filled={n_filled} empty={n_empty} spec_review={n_review}")
    print("primary distribution:")
    for k, v in sorted(primary_counts.items(), key=lambda kv: -kv[1]):
        print(f"  {v:5}  {k}")


if __name__ == "__main__":
    import sys

    models = load_models()
    print(f"loaded {len(models)} models")
    preview_only = "--preview" in sys.argv
    write_sheet(models, preview_only=preview_only)
